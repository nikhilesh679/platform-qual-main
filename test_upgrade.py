import pandas as pd
import pytest
from connect import connecting_node
from openpyxl import load_workbook
import re
import time
import subprocess


@pytest.mark.platform_qual
def test_run_upgrade():
    # Define the correct path to the platform-qual-main directory
    platform_qual_path = '/home/ubuntu/platform-qual-main'
    # Construct the command
    command = ['python3', 'pq.py', '-o', 'upgrade']

    # Run the command and capture output
    try:
        result = subprocess.run(
            command,
            cwd=platform_qual_path,
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=45*60  # 45 minutes
        )
        output = result.stdout.decode()
        # Assert the command executed successfully
        assert result.returncode == 0, f"Command failed with exit code {result.returncode}"
        output = result.stdout.decode()
        print("Output:\n", output)
    except subprocess.CalledProcessError as e:
        pytest.fail(f"Command execution failed with error: {e.stderr.decode()}")
    except FileNotFoundError:
        pytest.fail("The specified path or command was not found. Ensure the path and script are correct.")
        # Wait before proceeding to the SSH command
    #time.sleep(300)


@pytest.mark.platform_qual
def test_cluster_upgrade_status():
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'cluster bootstrap status'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster upgrade status"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster upgrade status', cleaned_output)
    ssh.close()



def update_excel(excel_file_path, target_scenario, cleaned_out):
    """
    Function to update the Excel file with the cleaned output.
    """
    try:
        # Load the workbook and worksheet
        wb = load_workbook(excel_file_path)
        ws = wb.active

        # Find columns and update the row
        scenario_col, result_col = None, None

        # Locate the columns by names
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == 'verification':
                scenario_col = col[0].column
            elif col[0].value == 'Actual result':
                result_col = col[0].column

        if scenario_col is None or result_col is None:
            raise ValueError("Required columns 'verification' or 'Actual result' not found in the Excel sheet.")

        for row in range(2, ws.max_row + 1):  # Skip the header row
            if ws.cell(row=row, column=scenario_col).value == target_scenario:
                ws.cell(row=row, column=result_col).value = cleaned_out
                wb.save(excel_file_path)
                wb.close()
                print(f"Output successfully updated in the Excel file at {excel_file_path}")
                return

        raise ValueError(f"The verification '{target_scenario}' was not found in the Excel sheet.")
    except Exception as e:
        print(f"An error occurred while updating the Excel: {e}") 




