import pandas as pd
import pytest
from connect import connecting_node
from openpyxl import load_workbook
import re
import time
import subprocess

@pytest.mark.platform_qual
def test_cluster_reset_node_force():
    """
    Test case to verify cluster reset_node force.
    """
    '''ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command="cluster reset_node force"
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster reset_node force"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('yes\n')
    stdin.write('\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('yes\n')
    stdin.write('\n')
    stdin.flush()
    time.sleep(500)
    stdin.write('\n')
    stdin.flush()'''
    ssh=connecting_node()
    new_command= "cd /tmp/sdtests && tail sdreset.out.txt"
    stdin, stdout, stderr = ssh.exec_command(new_command)
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{out}"
    print("Command Output:")
    print(full_output)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster reset_node force', cleaned_out)
    ssh.close()



def update_excel(excel_file_path, target_scenario, cleaned_out):
    """
    Function to update the Excel file with the cleaned output.
    """
    try:
        # Load the workbook and worksheet
        wb = load_workbook(excel_file_path)
        ws = wb.active

        # Find columns and update the row
        scenario_col, result_col = None, None

        # Locate the columns by names
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == 'verification':
                scenario_col = col[0].column
            elif col[0].value == 'Actual result':
                result_col = col[0].column

        if scenario_col is None or result_col is None:
            raise ValueError("Required columns 'verification' or 'Actual result' not found in the Excel sheet.")

        for row in range(2, ws.max_row + 1):  # Skip the header row
            if ws.cell(row=row, column=scenario_col).value == target_scenario:
                ws.cell(row=row, column=result_col).value = cleaned_out
                wb.save(excel_file_path)
                wb.close()
                print(f"Output successfully updated in the Excel file at {excel_file_path}")
                return

        raise ValueError(f"The verification '{target_scenario}' was not found in the Excel sheet.")
    except Exception as e:
        print(f"An error occurred while updating the Excel: {e}")

