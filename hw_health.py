import time
import pandas as pd
import pytest
import subprocess
import re
from openpyxl import load_workbook
from connect import connecting_node


def update_excel(excel_file_path, target_scenario, cleaned_out):
    """
    Function to update the Excel file with the cleaned output.
    """
    try:
        # Load the workbook and worksheet
        wb = load_workbook(excel_file_path)
        ws = wb.active

        # Find columns and update the row
        scenario_col, result_col = None, None

        # Locate the columns by names
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == 'verification':
                scenario_col = col[0].column
            elif col[0].value == 'Actual result':
                result_col = col[0].column

        if scenario_col is None or result_col is None:
            raise ValueError("Required columns 'verification' or 'Actual result' not found in the Excel sheet.")

        for row in range(2, ws.max_row + 1):  # Skip the header row
            if ws.cell(row=row, column=scenario_col).value == target_scenario:
                ws.cell(row=row, column=result_col).value = cleaned_out
                wb.save(excel_file_path)
                wb.close()
                print(f"Output successfully updated in the Excel file at {excel_file_path}")
                return

        raise ValueError(f"The verification '{target_scenario}' was not found in the Excel sheet.")
    except Exception as e:
        print(f"An error occurred while updating the Excel: {e}")





@pytest.mark.platform_qual
def test_hw_health():
    """
    Test case to verify cluster hw_health.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'cluster hw_health'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster hw_health"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster hw_health', cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_ipmi_related_info():
    """
    Test case to verify ipmi related info
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'sudo ipmitool mc info'
    full_command = "sudo ipmitool mc info"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'sudo ipmitool mc info', cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_ipmi_related_information():
    """
    Test case to verify ipmi related info
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "sudo ipmitool lan print 1"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'sudo ipmitool lan print 1', cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_reboot_node():
    """
    Test case to verify reboot node.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'cluster reboot node'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster reboot node"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('yes')
    stdin.flush()
    time.sleep(2)
    stdin.write('\n')
    stdin.flush()
    time.sleep(2)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster reboot node', cleaned_output)
    ssh.close()
