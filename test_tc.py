import time
import pandas as pd
import pytest
import subprocess
import re
from openpyxl import load_workbook
from connect import connecting_node
 
 
@pytest.mark.platform_qual
def test_run_pxe_manufacturing():
    """
    Test to execute the PXE manufacturing process using the pq.py script.
    """
    # Define the correct path to the platform-qual-main directory
    platform_qual_path = '/home/ubuntu/platform-qual-main'
 
    # Construct the command
    command = ['python3', 'pq.py', '-o', 'pxe']
 
    # Run the command and capture output
    try:
        result = subprocess.run(
            command,
            cwd=platform_qual_path,
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        # Assert the command executed successfully
        assert result.returncode == 0, f"Command failed with exit code {result.returncode}"
 
        # Check the stdout for expected content
        output = result.stdout.decode()
        assert "Welcome to Platform Qual testing" in output, "Output does not contain the welcome message"
 
        # Print output (optional for debugging)
        print("Output:\n", output)
 
    except subprocess.CalledProcessError as e:
        pytest.fail(f"Command execution failed with error: {e.stderr.decode()}")
 
    except FileNotFoundError:
        pytest.fail("The specified path or command was not found. Ensure the path and script are correct.")
 
    # Wait before proceeding to the SSH command
    time.sleep(2000)
 
    # Establish an SSH connection using the connecting_node function from connect module
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
 
    # Replace '/path/to/cluster' with the actual path to the `cluster` command
    full_command = f"/opt/rubrik/tools/rkcli_internal.sh cluster mfg_status"
    stdin, stdout, stderr = ssh.exec_command(full_command)
 
    # Read command output and errors
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
 
    # Output for debugging (optional)
    print("Command Output: ", out)
    print("Command Error: ", err)
 
    expected_output = "Manufacture successfully completed"
 
    # Check and print the expected output
    if expected_output in out:
        print(expected_output)
    assert expected_output in out, f"Expected output not found in command output. Actual output: {out}"
    assert err == "", f"Unexpected error: {err}"
    cleaned_out = re.sub(r'=+', '', out).strip()
    target_scenario = 'Perform PXE manufacturing by booting to USB boot with bios changes'
    # output_data = {
    #     'Testcase': ['Perform PXE manufacturing by booting to USB boot with bios changes'],
    #     'Command': ['cluster mfg_status'],
    #     'Actual result': [cleaned_out]
    # }
    # df = pd.DataFrame(output_data)
    #
    # # Print DataFrame to confirm content before writing to Excel
    # print("DataFrame Content:")
    # print(df)
    excel_file_path = '/home/ubuntu/platform-qual-main/automation_excel_file.xlsx'
    try:
        # Load the workbook and worksheet
        wb = load_workbook(excel_file_path)
        ws = wb.active
        # Find columns and update the row
        scenario_col, result_col = None, None
 
        # Locate the columns by names
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == 'Test Scenario':
                scenario_col = col[0].column
            elif col[0].value == 'Actual result':
                result_col = col[0].column
 
        if scenario_col is None or result_col is None:
            raise ValueError("Required columns 'Test Scenario' or 'Actual result' not found in the Excel sheet.")
        for row in range(2, ws.max_row + 1):  # Skip the header row
            if ws.cell(row=row, column=scenario_col).value == target_scenario:
                ws.cell(row=row, column=result_col).value = cleaned_out
                wb.save(excel_file_path)
                wb.close()
                print(f"Output successfully updated in the Excel file at {excel_file_path}")
                return
        raise ValueError(f"The test scenario '{target_scenario}' was not found in the Excel sheet.")
    except Exception as e:
        print(f"An error occurred while updating the Excel: {e}")
 
# Additional pytest configuration to avoid the warning (if required)
# Ensure you create a pytest.ini file in the root directory of your project with the following content:
 
# pytest.ini
# [pytest]
# markers =
# platform_qual: mark a test as a platform qualification test