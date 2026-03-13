import time
import pandas as pd
import pytest
import subprocess
import re
from openpyxl import load_workbook
from connect import connecting_node


def update_excel(excel_file_path, target_scenario, cleaned_out):
    """
    Function to update the Excel file with the cleaned output.
    """
    try:
        # Load the workbook and worksheet
        wb = load_workbook(excel_file_path)
        ws = wb.active

        # Find columns and update the row
        scenario_col, result_col = None, None

        # Locate the columns by names
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == 'verification':
                scenario_col = col[0].column
            elif col[0].value == 'Actual result':
                result_col = col[0].column

        if scenario_col is None or result_col is None:
            raise ValueError("Required columns 'verification' or 'Actual result' not found in the Excel sheet.")

        for row in range(2, ws.max_row + 1):  # Skip the header row
            if ws.cell(row=row, column=scenario_col).value == target_scenario:
                ws.cell(row=row, column=result_col).value = cleaned_out
                wb.save(excel_file_path)
                wb.close()
                print(f"Output successfully updated in the Excel file at {excel_file_path}")
                return

        raise ValueError(f"The verification '{target_scenario}' was not found in the Excel sheet.")
    except Exception as e:
        print(f"An error occurred while updating the Excel: {e}")


@pytest.mark.platform_qual
def test_network_route():
    """
    Test case to verify network route.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'network route'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network route"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network route', cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_network_hostname():
    """
    Test case to verify network hostname.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'network hostname'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network hostname"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network hostname', cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_network_hosts():
    """
    Test case to verify network hosts.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'network hosts'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network hosts"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network hosts', cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_network_ifconfig():
    """
    Test case to verify network ifconfig.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'network ifconfig'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network ifconfig"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network ifconfig', cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_network_ping():
    """
    Test case to verify network ping.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'network ping'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network ping google.com -c 4"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network ping <hostname>',
                 cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_network_set_default_gateway():
    """
    Test case to verify network set default gateway.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'network set default gateway'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network set_default_gateway"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('bond0\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('10.0.0.255')
    stdin.flush()
    time.sleep(2)
    stdin.write('\n')
    stdin.flush()
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network set_default_gateway',
                 cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_network_route_add():
    """
    Test case to verify network static route add.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'network static_route add'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network static_route add"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('10.0.0.0\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('255.255.0.0\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('N\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('bond0\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('10.0.0.255\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('yes')
    stdin.flush()
    time.sleep(2)
    stdin.write('\n')
    stdin.flush()

    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network static_route add',
                 cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_network_static_route_delete():
    """
    Test case to verify network static route delete.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'network static_route delete'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network static_route delete"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('1\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('yes')
    stdin.flush()
    time.sleep(2)
    stdin.write('\n')
    stdin.flush()
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network static_route delete',
                 cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_support_bundle():
    """
    Test case to verify support bundle.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'support cluster_support_bundle'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh support cluster_support_bundle"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('local')
    stdin.flush()
    time.sleep(2)
    stdin.write('\n')
    stdin.flush()
    time.sleep(30)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'support cluster_support_bundle',
                 cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_support_local_bundle():
    """
    Test case to verify support bundle.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'support local_support_bundle'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh support local_support_bundle"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('local')
    stdin.flush()
    time.sleep(2)
    stdin.write('\n')
    stdin.flush()
    time.sleep(30)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'support local_support_bundle',
                 cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_network_check_connectivity():
    """
    Test case to verify network check connectivity.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    get_hostname_command = "/opt/rubrik/tools/rkcli_internal.sh network hostname"
    print(f"Executing remote command to get hostname: '{get_hostname_command}'")
    stdin_hostname, stdout_hostname, stderr_hostname = ssh.exec_command(get_hostname_command)
    captured_hostname = stdout_hostname.read().decode().strip()
    hostname_error = stderr_hostname.read().decode().strip()

    if hostname_error:
        pytest.fail(f"Error getting hostname from remote node: {hostname_error}")
    if not captured_hostname:
        pytest.fail("Hostname command returned empty output.")

    print(f"Successfully captured hostname: '{captured_hostname}'")
    target_host_for_check = captured_hostname
    target_port_for_check = "22"
    command = 'network check_connectivity'
    full_command = f"/opt/rubrik/tools/rkcli_internal.sh network check_connectivity {target_host_for_check} {target_port_for_check}"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx','network check_connectivity <host> <port>', cleaned_output)
    ssh.close()


@pytest.mark.platform_qual
def test_support_log_view():
    """
    Test case to verify support bundle.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    command = 'support log_view'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh support log_view"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('exit')
    stdin.flush()
    time.sleep(2)
    stdin.write('\n')
    stdin.flush()
    time.sleep(5)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'support log_view', cleaned_output)
    ssh.close()

'''
@pytest.mark.platform_qual
def test_network_re_ip():
    """
    Test case to verify network re_ip.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    get_hostname_command = "/opt/rubrik/tools/rkcli_internal.sh network hostname"
    print(f"Executing remote command to get hostname: '{get_hostname_command}'")
    stdin_h, stdout_h, stderr_h = ssh.exec_command(get_hostname_command)
    current_node_hostname = stdout_h.read().decode().strip()
    if not current_node_hostname:
        pytest.fail("Failed to retrieve current node's hostname.")
    print(f"Current node hostname: '{current_node_hostname}'")

    get_node_table_command = "/opt/rubrik/tools/rkcli_internal.sh cluster node_table"
    print(f"Executing remote command: '{get_node_table_command}'")
    stdin_nt, stdout_nt, stderr_nt = ssh.exec_command(get_node_table_command)
    node_table_output = stdout_nt.read().decode().strip()
    node_table_error = stderr_nt.read().decode().strip()

    if node_table_error:
        print(f"Error getting node table: {node_table_error}")
        pytest.fail(f"Failed to retrieve node table: {node_table_error}")

    print(f"Node table output received:\n{node_table_output}")

    management_ip_address = None

    lines = node_table_output.split('\n')
    found_node_section = False
    for line_num, line in enumerate(lines):  
        line = line.strip()
        # Look for the node's section based on hostname/node ID
        if f"node id: {current_node_hostname}" in line or f"hostname: {current_node_hostname}" in line:
            found_node_section = True


        if found_node_section and "management IP address:" in line:
    # Extract the IP address after "management IP address:"
            ip_match = re.search(r'management IP address:\s*(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
        if ip_match:
            management_ip_address = ip_match.group(1)
            break  

    if not management_ip_address:
         pytest.fail(f"Could not find management IP address for node '{current_node_hostname}' in node table output.")

    print(f"Extracted dynamic management IP: '{management_ip_address}'")

    command = 'network re_ip'
    full_command = "/opt/rubrik/tools/rkcli_internal.sh network re_ip"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('10.0.0.255\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('255.255.0.0\n')
    stdin.flush()
    time.sleep(2)
    stdin.write('\n')
    stdin.flush()
    stdin.write('\n')
    stdin.flush()
    stdin.write(f'{management_ip_address}\n')
    stdin.flush()
    time.sleep(5)
    stdin.write('Yes')
    stdin.flush()
    time.sleep(5)
    stdin.write('\n')
    stdin.flush()
    time.sleep(120)

    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    # print("Command Error:")
    print(err)
    # Clean the output
    cleaned_output = re.sub(r'=+', '', full_output).strip()

    # Update Excel file

    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'network re_ip',
                 cleaned_output)
    ssh.close()
'''








    
