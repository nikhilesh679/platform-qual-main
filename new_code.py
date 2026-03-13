import time
import pandas as pd
import pytest
import subprocess
import re
from openpyxl import load_workbook
from connect import connecting_node
@pytest.mark.platform_qual
def test_run_pxe_manufacturing():
    """
    Test to execute the PXE manufacturing process using the pq.py script.
    """
    # Define the correct path to the platform-qual-main directory
    platform_qual_path = '/home/ubuntu/platform-qual-main'
    # Construct the command
    command = ['python3', 'pq.py', '-o', 'pxe']
    # Run the command and capture output
    try:
        result = subprocess.run(
            command,
            cwd=platform_qual_path,
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        # Assert the command executed successfully
        assert result.returncode == 0, f"Command failed with exit code {result.returncode}"
        # Check the stdout for expected content
        output = result.stdout.decode()
        assert "Welcome to Platform Qual testing" in output, "Output does not contain the welcome message"
        # Print output (optional for debugging)
        print("Output:\n", output)
    except subprocess.CalledProcessError as e:
        pytest.fail(f"Command execution failed with error: {e.stderr.decode()}")
    except FileNotFoundError:
        pytest.fail("The specified path or command was not found. Ensure the path and script are correct.")
    # Wait before proceeding to the SSH command
    time.sleep(2000)
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster mfg_status"
    expected_output = "Manufacture successfully completed"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    assert expected_output in out, f"Expected output not found in command output. Actual output: {out}"
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster mfg_status', cleaned_out)

'''@pytest.mark.platform_qual
def test_cluster_discover():
    """
    Test case to verify cluster discovery.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster discover"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx',
                 "cluster discover (after removing the '/var/lib/rubrik/hosts_to_ipv6_mapping' file from node)",
                 cleaned_out)
@pytest.mark.platform_qual
def test_cluster_reset_node_status():
    """
    Test case to verify cluster reset_node_status.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster reset_node_status"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster reset_node_status',
                 cleaned_out)
@pytest.mark.platform_qual
def test_cluster_reset_node_force():
    """
    Test case to verify cluster reset_node force.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster reset_node force"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('yes\n')
    stdin.flush()
    stdin.write('yes\n')
    stdin.flush()
    time.sleep(200)
    stdin.write('\n')
    stdin.flush()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster reset_node force', cleaned_out)
@pytest.mark.platform_qual
def test_cluster_reset_node_skip_ipmi_network_reset():
    """
    Test case to verify cluster reset_node force.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster reset node skip_ipmi_network_reset"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('yes\n')
    stdin.flush()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster reset node skip_ipmi_network_reset',
                 cleaned_out)

@pytest.mark.platform_qual
def test_cluster_reset_node_keep_logs():
    """
    Test case to verify cluster reset_node force.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster reset_node keep_logs"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    stdin.write('yes\n')
    stdin.flush()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster reset_node keep_logs',
                 cleaned_out)


# import subprocess
# import pytest
# import re
# import time

@pytest.mark.platform_qual
def test_run_bootstrap():
    """
    Test to execute the bootstrap process using the pq.py script.
    """
    # Define the correct path to the platform-qual-main directory
    platform_qual_path = '/home/ubuntu/platform-qual-main'
    # Construct the command
    command = ['python3', 'pq.py', '-o', 'bootstrap']

    # Run the command and capture output
    try:
        result = subprocess.run(
            command,
            cwd=platform_qual_path,
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=45*60  # 45 minutes
        )
        output = result.stdout.decode()
        # Assert the command executed successfully
        assert result.returncode == 0, f"Command failed with exit code {result.returncode}"
        # Check the stdout for expected content
        if "System setup is complete." in output:
            output_message = "Bootstrap completed successfully"
        else:
            output_message = "Bootstrap did not complete as expected"
        # Print output (optional for debugging)
        print("Output:\n", output)
    except subprocess.TimeoutExpired:
        pytest.fail("Command execution timed out after 45 minutes")
    except subprocess.CalledProcessError as e:
        output_message = f"Command execution failed with error: {e.stderr.decode()}"
    except FileNotFoundError:
        pytest.fail("The specified path or command was not found. Ensure the path and script are correct.")

    # Clean the output
    cleaned_output = re.sub(r'=+', '', output).strip()

    # Update Excel file
    update_excel('/home/ubuntu/platform-qual-main/automation.xlsx', 'cluster Bootstrap', output_message)
@pytest.mark.platform_qual
def test_cluster_get_node_statuses():
    """
    Test case to verify cluster reset_node_status.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster get_node_statuses"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster get_node_statuses',
                 cleaned_out)
@pytest.mark.platform_qual
def test_cluster_reset_node_status():
    """
    Test case to verify cluster reset_node_status.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster reset_node_status"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster reset_node_status',
                 cleaned_out)
#cluster cluster_uuid
@pytest.mark.platform_qual
def test_cluster_cluster_uuid():
    """
    Test case to verify cluster reset_node_status.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster cluster_uuid"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster cluster_uuid',
                 cleaned_out)
#cluster cluster_name
@pytest.mark.platform_qual
def test_cluster_cluster_name():
    """
    Test case to verify cluster reset_node_status.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster cluster_name"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster cluster_name',
                 cleaned_out)
#cluster node_table
@pytest.mark.platform_qual
def test_cluster_node_table():
    """
    Test case to verify cluster reset_node_status.
    """
    ssh = connecting_node()
    if ssh is None:
        pytest.fail("Failed to connect to the remote node.")
    else:
        print("Successfully connected to node")
    full_command = "/opt/rubrik/tools/rkcli_internal.sh cluster node_table"
    stdin, stdout, stderr = ssh.exec_command(full_command)
    # out = stdout.read().decode().strip()
    # err = stderr.read().decode().strip()
    command_input = f"{full_command}\n"
    out = stdout.read().decode().strip()
    err = stderr.read().decode().strip()
    full_output = f"{command_input}{out}"
    print("Command Output:")
    print(out)
    print("Command Error:")
    print(err)
    cleaned_out = re.sub(r'=+', '', full_output).strip()
    update_excel('/home/ubuntu/platform-qual-main/automation_excel_file.xlsx', 'cluster node_table',
                 cleaned_out)'''




def update_excel(excel_file_path, target_scenario, cleaned_out):
    """
    Function to update the Excel file with the cleaned output.
    """
    try:
        # Load the workbook and worksheet
        wb = load_workbook(excel_file_path)
        ws = wb.active

        # Find columns and update the row
        scenario_col, result_col = None, None

        # Locate the columns by names
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            if col[0].value == 'verification':
                scenario_col = col[0].column
            elif col[0].value == 'Actual result':
                result_col = col[0].column

        if scenario_col is None or result_col is None:
            raise ValueError("Required columns 'verification' or 'Actual result' not found in the Excel sheet.")

        for row in range(2, ws.max_row + 1):  # Skip the header row
            if ws.cell(row=row, column=scenario_col).value == target_scenario:
                ws.cell(row=row, column=result_col).value = cleaned_out
                wb.save(excel_file_path)
                wb.close()
                print(f"Output successfully updated in the Excel file at {excel_file_path}")
                return

        raise ValueError(f"The verification '{target_scenario}' was not found in the Excel sheet.")
    except Exception as e:
        print(f"An error occurred while updating the Excel: {e}")
